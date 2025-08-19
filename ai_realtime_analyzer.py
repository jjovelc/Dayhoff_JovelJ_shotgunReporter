#!/usr/bin/env python3
"""
Real-time AI Analyzer for Microbiome Plots
Generates personalized explanations based on actual plot data values
"""

import json
import requests
import pandas as pd
import numpy as np
import os
from typing import Dict, Any, Optional, List, Tuple
import sys
import openpyxl

class RealTimeMicrobiomeAnalyzer:
    def __init__(self, ollama_url: str = "http://localhost:11434", model: str = "gpt-oss:20b"):
        self.ollama_url = ollama_url
        self.model = model
        self.api_url = f"{ollama_url}/api/generate"
        
        # Data file paths
        self.data_dir = "."
        self.metadata_file = "metadata.tsv"
        self.level_files = {
            2: "all_child-UC_kraken2_250616_level_2.tsv",
            3: "all_child-UC_kraken2_250616_level_3.tsv", 
            4: "all_child-UC_kraken2_250616_level_4.tsv",
            5: "all_child-UC_kraken2_250616_level_5.tsv",
            6: "all_child-UC_kraken2_250616_level_6.tsv",
            7: "all_child-UC_kraken2_250616_level_7.tsv"
        }
    
    def extract_alpha_diversity_data(self, taxonomic_level: str) -> Dict[str, float]:
        """
        Extract actual alpha diversity values from the data files
        """
        try:
            # Map taxonomic level to file level
            level_mapping = {
                "phylum": 2, "class": 3, "order": 4, 
                "family": 5, "genus": 6, "species": 7
            }
            
            level = level_mapping.get(taxonomic_level.lower(), 2)
            level_file = self.level_files.get(level)
            
            if not level_file or not os.path.exists(level_file):
                return {}
            
            # Read the data file
            df = pd.read_csv(level_file, sep='\t', index_col=0)
            
            # Calculate Shannon diversity for each sample
            diversity_data = {}
            for sample in df.columns:
                # Get abundances for this sample
                abundances = df[sample].values
                # Remove zeros and calculate Shannon diversity
                non_zero = abundances[abundances > 0]
                if len(non_zero) > 0:
                    # Normalize to proportions
                    proportions = non_zero / non_zero.sum()
                    # Calculate Shannon diversity: -sum(p * log(p))
                    shannon = -sum(p * (np.log(p) if p > 0 else 0) for p in proportions)
                    diversity_data[sample] = shannon
            
            return diversity_data
            
        except Exception as e:
            print(f"Error extracting alpha diversity data: {e}")
            return {}
    
    def extract_stacked_barplot_data(self, taxonomic_level: str) -> Tuple[List[str], Dict[str, List[float]]]:
        """
        Extract actual abundance data for stacked barplots
        """
        try:
            # Map taxonomic level to file level
            level_mapping = {
                "phylum": 2, "class": 3, "order": 4, 
                "family": 5, "genus": 6, "species": 7
            }
            
            level = level_mapping.get(taxonomic_level.lower(), 2)
            level_file = self.level_files.get(level)
            
            if not level_file or not os.path.exists(level_file):
                return [], {}
            
            # Read the data file
            df = pd.read_csv(level_file, sep='\t', index_col=0)
            
            # Get top 20 taxa by total abundance
            total_abundance = df.sum(axis=1)
            top_taxa = total_abundance.nlargest(20).index.tolist()
            
            # Get abundances for each sample
            abundances = {}
            for sample in df.columns:
                abundances[sample] = df.loc[top_taxa, sample].tolist()
            
            return top_taxa, abundances
            
        except Exception as e:
            print(f"Error extracting stacked barplot data: {e}")
            return [], {}
    
    def extract_pcoa_data(self, taxonomic_level: str) -> Dict[str, List[float]]:
        """
        Extract actual abundance data for PCoA analysis
        """
        try:
            # Map taxonomic level to file level
            level_mapping = {
                "phylum": 2, "class": 3, "order": 4, 
                "family": 5, "genus": 6, "species": 7
            }
            
            level = level_mapping.get(taxonomic_level.lower(), 2)
            level_file = self.level_files.get(level)
            
            if not level_file or not os.path.exists(level_file):
                return {}
            
            # Read the data file
            df = pd.read_csv(level_file, sep='\t', index_col=0)
            
            # Get abundances for each sample
            abundances = {}
            for sample in df.columns:
                abundances[sample] = df[sample].tolist()
            
            return abundances
            
        except Exception as e:
            print(f"Error extracting PCoA data: {e}")
            return {}
    
    def get_sample_metadata(self) -> Dict[str, str]:
        """
        Get sample metadata (Control vs UC)
        """
        try:
            if not os.path.exists(self.metadata_file):
                return {}
            
            df = pd.read_csv(self.metadata_file, sep='\t')
            metadata = {}
            
            for _, row in df.iterrows():
                sample_id = row.get('sample', '')
                condition = row.get('group', '')
                if sample_id and condition:
                    metadata[sample_id] = condition
            
            return metadata
            
        except Exception as e:
            print(f"Error reading metadata: {e}")
            return {}
    
    def calculate_taxa_control_uc_ratios(self, taxonomic_level: str) -> List[Dict[str, any]]:
        """
        Calculate Control/UC ratios for all taxa at a given taxonomic level
        
        Args:
            taxonomic_level: Taxonomic level (can be "phylum", "class", "order", "family", "genus", "species" 
                             or numeric levels "2", "3", "4", "5", "6", "7")
            
        Returns:
            List of dictionaries with taxon name and Control/UC ratio
        """
        try:
            # Map taxonomic level to file level (handle both string and numeric inputs)
            level_mapping = {
                # String inputs
                "phylum": 2, "class": 3, "order": 4, 
                "family": 5, "genus": 6, "species": 7,
                # Numeric inputs
                "2": 2, "3": 3, "4": 4, "5": 5, "6": 6, "7": 7
            }
            
            level = level_mapping.get(taxonomic_level.lower(), 2)
            level_file = self.level_files.get(level)
            
            if not level_file or not os.path.exists(level_file):
                return []
            
            # Create mapping structure that relates each taxonomic file with its expected prefix
            file_prefix_mapping = {
                2: 'p__',  # phylum file -> only include taxa starting with p__
                3: 'c__',  # class file -> only include taxa starting with c__
                4: 'o__',  # order file -> only include taxa starting with o__
                5: 'f__',  # family file -> only include taxa starting with f__
                6: 'g__',  # genus file -> only include taxa starting with g__
                7: 's__'   # species file -> only include taxa starting with s__
            }
            
            expected_prefix = file_prefix_mapping[level]
            
            # Read the data file
            df = pd.read_csv(level_file, sep='\t', index_col=0)
            
            # Get sample metadata
            metadata = self.get_sample_metadata()
            
            # Separate control vs UC samples
            control_samples = []
            uc_samples = []
            for sample in df.columns:
                condition = metadata.get(sample, 'Unknown')
                if 'control' in condition.lower():
                    control_samples.append(sample)
                else:
                    uc_samples.append(sample)
            
            # Calculate ratios for each taxon
            taxa_ratios = []
            filtered_count = 0
            
            for taxon in df.index:
                # Extract the family-level designation from within the full taxonomic path
                # The taxon string format is: k__Kingdom|p__Phylum|c__Class|o__Order|f__Family|g__Genus|s__Species
                taxon_parts = taxon.split('|')
                
                # Find the part that corresponds to our level
                level_prefix = {
                    2: 'p__',  # phylum
                    3: 'c__',  # class
                    4: 'o__',  # order
                    5: 'f__',  # family
                    6: 'g__',  # genus
                    7: 's__'   # species
                }[level]
                
                # Look for the part that starts with our expected prefix
                level_taxon = None
                for part in taxon_parts:
                    if part.startswith(level_prefix):
                        level_taxon = part
                        break
                
                # Only include taxa that have the expected taxonomic level designation
                if not level_taxon:
                    continue  # Skip taxa that don't have the requested level
                
                filtered_count += 1
                
                # Extract the taxon name at this level (remove prefix and clean up)
                taxon_name = level_taxon[len(level_prefix):]  # Remove the prefix
                clean_taxon_name = taxon_name.replace('_', ' ')  # Replace underscores with spaces
                
                # Get abundances for this taxon
                control_values = [df.loc[taxon, sample] for sample in control_samples if sample in df.columns]
                uc_values = [df.loc[taxon, sample] for sample in uc_samples if sample in df.columns]
                
                # Calculate averages
                control_avg = sum(control_values) / len(control_values) if control_values else 0
                uc_avg = sum(uc_values) / len(uc_values) if uc_values else 0
                
                # Calculate standard deviation for UC samples
                uc_std = 0
                if len(uc_values) > 1:
                    uc_mean = uc_avg
                    variance = sum((x - uc_mean) ** 2 for x in uc_values) / (len(uc_values) - 1)
                    uc_std = variance ** 0.5
                
                # Calculate ratio (Control/UC)
                if uc_avg > 0:
                    ratio = control_avg / uc_avg
                else:
                    ratio = float('inf') if control_avg > 0 else 0
                
                # Add to results
                taxa_ratios.append({
                    'taxon': f"{level_prefix}{clean_taxon_name}",  # Keep prefix for verification
                    'control_avg': control_avg,
                    'uc_avg': uc_avg,
                    'uc_std': uc_std,  # Add standard deviation
                    'control_uc_ratio': ratio,
                    'control_samples_count': len(control_values),
                    'uc_samples_count': len(uc_values)
                })
            
            # Sort by ratio (highest to lowest)
            taxa_ratios.sort(key=lambda x: x['control_uc_ratio'], reverse=True)
            
            return taxa_ratios
            
        except Exception as e:
            print(f"Error calculating taxa ratios: {e}")
            return []
    
    def generate_taxa_comparison_tsv(self, taxonomic_level: str) -> str:
        """
        Generate TSV content for taxa comparison table
        
        Args:
            taxonomic_level: Taxonomic level to analyze
            
        Returns:
            TSV string content
        """
        try:
            taxa_data = self.calculate_taxa_control_uc_ratios(taxonomic_level)
            
            if not taxa_data:
                return "No data available for the specified taxonomic level."
            
            # Create TSV header
            tsv_content = "Taxon\tControl_Average\tUC_Average\tUC_StdDev\tControl_UC_Ratio\tControl_Samples\tUC_Samples\n"
            
            # Add data rows
            for taxon_info in taxa_data:
                tsv_content += f"{taxon_info['taxon']}\t"
                tsv_content += f"{taxon_info['control_avg']:.6f}\t"
                tsv_content += f"{taxon_info['uc_avg']:.6f}\t"
                tsv_content += f"{taxon_info['uc_std']:.6f}\t"
                tsv_content += f"{taxon_info['control_uc_ratio']:.6f}\t"
                tsv_content += f"{taxon_info['control_samples_count']}\t"
                tsv_content += f"{taxon_info['uc_samples_count']}\n"
            
            return tsv_content
            
        except Exception as e:
            print(f"Error generating TSV: {e}")
            return f"Error generating TSV: {e}"
    
    def generate_taxa_comparison_excel(self, taxonomic_level: str) -> bytes:
        """
        Generate Excel file for taxa comparison table
        
        Args:
            taxonomic_level: Taxonomic level to analyze
            
        Returns:
            Excel file as bytes
        """
        try:
            taxa_data = self.calculate_taxa_control_uc_ratios(taxonomic_level)
            
            if not taxa_data:
                # Create empty Excel with error message
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Taxa Comparison"
                ws['A1'] = "No data available for the specified taxonomic level."
                return wb.save()
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Taxa Comparison - {taxonomic_level.title()}"
            
            # Add headers
            headers = ["Taxon", "Control_Average", "UC_Average", "UC_StdDev", "Control_UC_Ratio", "Control_Samples", "UC_Samples"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                ws.cell(row=1, column=col).fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row, taxon_info in enumerate(taxa_data, 2):
                ws.cell(row=row, column=1, value=taxon_info['taxon'])
                ws.cell(row=row, column=2, value=round(taxon_info['control_avg'], 6))
                ws.cell(row=row, column=3, value=round(taxon_info['uc_avg'], 6))
                ws.cell(row=row, column=4, value=round(taxon_info['uc_std'], 6))
                ws.cell(row=row, column=5, value=round(taxon_info['control_uc_ratio'], 6))
                ws.cell(row=row, column=6, value=taxon_info['control_samples_count'])
                ws.cell(row=row, column=7, value=taxon_info['uc_samples_count'])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            print(f"Error generating Excel: {e}")
            # Return minimal Excel with error
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = f"Error generating Excel: {e}"
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

    def analyze_taxon_plot(self, taxon_name: str, sample_data: Dict[str, float], 
                          sample_names: list, rpm_values: list) -> str:
        """
        Generate personalized AI analysis for a specific taxon plot
        """
        # Create a detailed prompt with the actual data
        prompt = self._create_taxon_analysis_prompt(taxon_name, sample_data, sample_names, rpm_values)
        
        try:
            response = self._call_ollama(prompt)
            return response
        except Exception as e:
            return self._generate_fallback_analysis(taxon_name, sample_data, sample_names, rpm_values)
    
    def analyze_diversity_plot(self, plot_type: str, diversity_data: Dict[str, float]) -> str:
        """
        Generate personalized AI analysis for diversity plots
        """
        prompt = self._create_diversity_analysis_prompt(plot_type, diversity_data)
        
        try:
            response = self._call_ollama(prompt)
            return response
        except Exception as e:
            return self._generate_fallback_diversity_analysis(plot_type, diversity_data)
    
    def analyze_stacked_plot(self, plot_type: str, top_taxa: list, abundances: Dict[str, list]) -> str:
        """
        Generate personalized AI analysis for stacked bar plots
        """
        prompt = self._create_stacked_analysis_prompt(plot_type, top_taxa, abundances)
        
        try:
            response = self._call_ollama(prompt)
            return response
        except Exception as e:
            return self._generate_fallback_stacked_analysis(plot_type, top_taxa, abundances)
    
    def analyze_pcoa_plot(self, plot_type: str, abundances: Dict[str, list]) -> str:
        """
        Generate personalized AI analysis for PCoA plots
        """
        prompt = self._create_pcoa_analysis_prompt(plot_type, abundances)
        
        try:
            response = self._call_ollama(prompt)
            return response
        except Exception as e:
            return self._generate_fallback_pcoa_analysis(plot_type, abundances)
    
    def _create_taxon_analysis_prompt(self, taxon_name: str, sample_data: Dict[str, float], 
                                    sample_names: list, rpm_values: list) -> str:
        """
        Create a detailed prompt for taxon-specific analysis
        """
        # Find control vs UC samples
        control_samples = [name for name in sample_names if 'ctrl' in name.lower() or 'control' in name.lower()]
        uc_samples = [name for name in sample_names if name not in control_samples]
        
        # Calculate statistics
        control_values = [sample_data.get(name, 0) for name in control_samples if name in sample_data]
        uc_values = [sample_data.get(name, 0) for name in uc_samples if name in sample_data]
        
        control_avg = sum(control_values) / len(control_values) if control_values else 0
        uc_avg = sum(uc_values) / len(uc_values) if uc_values else 0
        
        max_value = max(rpm_values) if rpm_values else 0
        min_value = min(rpm_values) if rpm_values else 0
        
        prompt = f"""
You are a microbiome expert explaining results to patients in simple terms. Analyze this specific data:

TAXON: {taxon_name}
SAMPLE DATA: {json.dumps(sample_data, indent=2)}
SAMPLE NAMES: {sample_names}
RPM VALUES: {rpm_values}

CONTROL SAMPLES: {control_samples} (Average: {control_avg:.3f} RPM)
UC SAMPLES: {uc_samples} (Average: {uc_avg:.3f} RPM)
RANGE: {min_value:.3f} to {max_value:.3f} RPM

Provide a personalized analysis that:
1. Mentions the specific taxon name
2. References actual RPM values and sample names
3. Compares control vs UC samples with specific numbers
4. Explains what the differences might mean in simple terms
5. Gives actionable insights for patients

Keep it under 150 words, simple language, and focus on the specific data provided.
"""
        return prompt
    
    def _create_diversity_analysis_prompt(self, plot_type: str, diversity_data: Dict[str, float]) -> str:
        """
        Create a detailed prompt for diversity plot analysis
        """
        # Calculate statistics
        values = list(diversity_data.values())
        avg_diversity = sum(values) / len(values) if values else 0
        max_diversity = max(values) if values else 0
        min_diversity = min(values) if values else 0
        
        # Get sample metadata
        metadata = self.get_sample_metadata()
        
        # Separate control vs UC samples
        control_samples = []
        uc_samples = []
        for sample, diversity in diversity_data.items():
            condition = metadata.get(sample, 'Unknown')
            if 'control' in condition.lower():
                control_samples.append((sample, diversity))
            else:
                uc_samples.append((sample, diversity))
        
        control_avg = sum(d for _, d in control_samples) / len(control_samples) if control_samples else 0
        uc_avg = sum(d for _, d in uc_samples) / len(uc_samples) if uc_samples else 0
        
        prompt = f"""
You are a microbiome expert explaining diversity results to patients in simple terms. Analyze this specific data:

PLOT TYPE: {plot_type}
DIVERSITY DATA: {json.dumps(diversity_data, indent=2)}
AVERAGE DIVERSITY: {avg_diversity:.3f}
RANGE: {min_diversity:.3f} to {max_diversity:.3f}

CONTROL SAMPLES: {[s[0] for s in control_samples]} (Average: {control_avg:.3f})
UC SAMPLES: {[s[0] for s in uc_samples]} (Average: {uc_avg:.3f})

Provide a personalized analysis that:
1. Explains what {plot_type} diversity means
2. References actual diversity values and sample names
3. Compares control vs UC samples with specific numbers
4. Explains what the differences might mean for gut health
5. Stress the importance of a diverse microbiome

Keep it under 200 words, simple language, and focus on the specific data provided.
"""
        return prompt
    
    def _create_stacked_analysis_prompt(self, plot_type: str, top_taxa: list, abundances: Dict[str, list]) -> str:
        """
        Create a detailed prompt for stacked bar plot analysis
        """
        # Get sample metadata
        metadata = self.get_sample_metadata()
        
        # Calculate total abundances for each taxon
        taxon_totals = {}
        for i, taxon in enumerate(top_taxa):
            total = sum(abundances[sample][i] for sample in abundances if i < len(abundances[sample]))
            taxon_totals[taxon] = total
        
        # Get top 5 most abundant taxa
        top_5_taxa = sorted(taxon_totals.items(), key=lambda x: x[1], reverse=True)[:5]
        
        prompt = f"""
You are a microbiome expert explaining microbiome composition results to patients in simple terms. Analyze this specific data:

PLOT TYPE: {plot_type}
TOP TAXA: {top_taxa}
TOP 5 MOST ABUNDANT: {top_5_taxa}
SAMPLE ABUNDANCES: {json.dumps(abundances, indent=2)}

Provide a personalized analysis that:
1. Explains what this {plot_type} plot shows
2. Mentions specific taxa names from the data
3. References actual abundance values
4. Explains what the patterns might mean for gut health
5. Establish a brief comparisson of taxa abundance between control and UC samples

Keep it under 200 words, simple language, and focus on the specific data provided.
"""
        return prompt
    
    def _create_pcoa_analysis_prompt(self, plot_type: str, abundances: Dict[str, list]) -> str:
        """
        Create a detailed prompt for PCoA plot analysis
        """
        # Get sample metadata
        metadata = self.get_sample_metadata()
        
        # Separate control vs UC samples
        control_samples = []
        uc_samples = []
        for sample in abundances.keys():
            condition = metadata.get(sample, 'Unknown')
            if 'control' in condition.lower():
                control_samples.append(sample)
            else:
                uc_samples.append(sample)
        
        prompt = f"""
You are a microbiome expert explaining PCoA results to patients in simple terms. Analyze this specific data:

PLOT TYPE: {plot_type}
CONTROL SAMPLES: {control_samples}
UC SAMPLES: {uc_samples}
ABUNDANCE DATA: {json.dumps(abundances, indent=2)}

Provide a personalized analysis that:
1. Explains what this PCoA plot shows
2. References the specific samples and their groups
3. Explains what the patterns might mean for gut health
4. Gives actionable insights for patients

Keep it under 150 words, simple language, and focus on the specific data provided.
"""
        return prompt
    
    def _call_ollama(self, prompt: str) -> str:
        """
        Call the Ollama API to generate AI analysis
        """
        payload = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
            "options": {
                "temperature": 0.7,
                "top_p": 0.9,
                "max_tokens": 300
            }
        }
        
        response = requests.post(self.api_url, json=payload, timeout=30)
        response.raise_for_status()
        
        result = response.json()
        return result.get('response', '').strip()
    
    def _generate_fallback_analysis(self, taxon_name: str, sample_data: Dict[str, float], 
                                  sample_names: list, rpm_values: list) -> str:
        """
        Generate a fallback analysis when AI is unavailable
        """
        control_samples = [name for name in sample_names if 'ctrl' in name.lower() or 'control' in name.lower()]
        uc_samples = [name for name in sample_names if name not in control_samples]
        
        control_values = [sample_data.get(name, 0) for name in control_samples if name in sample_data]
        uc_values = [sample_data.get(name, 0) for name in uc_samples if name in sample_data]
        
        control_avg = sum(control_values) / len(control_values) if control_values else 0
        uc_avg = sum(uc_values) / len(uc_values) if uc_values else 0
        
        return f"""This plot shows {taxon_name} abundance across your samples. 

Control samples show an average of {control_avg:.3f} RPM, while UC samples show {uc_avg:.3f} RPM. 

The difference suggests {taxon_name} may be {'higher' if uc_avg > control_avg else 'lower'} in your condition compared to healthy controls. This could indicate changes in your gut microbiome that may be related to your health status.

Consult with your healthcare provider about what these specific levels mean for your individual case."""
    
    def _generate_fallback_diversity_analysis(self, plot_type: str, diversity_data: Dict[str, float]) -> str:
        """
        Generate a fallback analysis for diversity plots
        """
        values = list(diversity_data.values())
        avg_diversity = sum(values) / len(values) if values else 0
        
        # Get sample metadata
        metadata = self.get_sample_metadata()
        
        # Separate control vs UC samples
        control_samples = []
        uc_samples = []
        for sample, diversity in diversity_data.items():
            condition = metadata.get(sample, 'Unknown')
            if 'control' in condition.lower():
                control_samples.append((sample, diversity))
            else:
                uc_samples.append((sample, diversity))
        
        control_avg = sum(d for _, d in control_samples) / len(control_samples) if control_samples else 0
        uc_avg = sum(d for _, d in uc_samples) / len(uc_samples) if uc_samples else 0
        
        return f"""This {plot_type} diversity plot shows how diverse your gut microbiome is across samples.

Your average diversity is {avg_diversity:.3f}. {'Higher diversity generally indicates a healthier, more balanced microbiome.' if avg_diversity > 3.0 else 'Lower diversity may suggest your microbiome needs support to become more balanced.'}

Control samples show an average diversity of {control_avg:.3f}, while UC samples show {uc_avg:.3f}. This comparison helps identify how your condition may affect microbiome diversity.

The specific values for each sample help identify which areas of your gut may need attention. Discuss these results with your healthcare provider for personalized recommendations."""
    
    def _generate_fallback_stacked_analysis(self, plot_type: str, top_taxa: list, abundances: Dict[str, list]) -> str:
        """
        Generate a fallback analysis for stacked plots
        """
        # Calculate total abundances for each taxon
        taxon_totals = {}
        for i, taxon in enumerate(top_taxa):
            total = sum(abundances[sample][i] for sample in abundances if i < len(abundances[sample]))
            taxon_totals[taxon] = total
        
        # Get top 5 most abundant taxa
        top_5_taxa = sorted(taxon_totals.items(), key=lambda x: x[1], reverse=True)[:5]
        
        return f"""This {plot_type} plot shows the composition of your gut microbiome, highlighting the most abundant bacteria groups.

The top taxa in your samples include: {', '.join([t[0] for t in top_5_taxa])}.

This visualization helps identify which bacteria are dominant in your gut and how they compare between samples. The patterns can reveal important information about your microbiome balance and health status.

Discuss these specific results with your healthcare provider for personalized insights."""
    
    def _generate_fallback_pcoa_analysis(self, plot_type: str, abundances: Dict[str, list]) -> str:
        """
        Generate a fallback analysis for PCoA plots
        """
        # Get sample metadata
        metadata = self.get_sample_metadata()
        
        # Separate control vs UC samples
        control_samples = []
        uc_samples = []
        for sample in abundances.keys():
            condition = metadata.get(sample, 'Unknown')
            if 'control' in condition.lower():
                control_samples.append(sample)
            else:
                uc_samples.append(sample)
        
        return f"""This PCoA plot shows how similar or different your microbiome is compared to others.

Control samples: {', '.join(control_samples)}
UC samples: {', '.join(uc_samples)}

This visualization helps identify if people with the same health condition have similar gut microbiomes, which could help develop treatments or understand how the disease affects the gut.

The plot shows the relationships between samples based on their microbial composition, helping to identify patterns that may be related to health status."""

    def generate_ai_summary(self, taxonomic_level: str, report_type: str = "technical") -> str:
        """
        Generate AI summary report using gpt-oss-20b
        
        Args:
            taxonomic_level: Taxonomic level to analyze
            report_type: "technical" or "lay" for different writing styles
            
        Returns:
            Generated summary text
        """
        try:
            taxa_data = self.calculate_taxa_control_uc_ratios(taxonomic_level)
            
            if not taxa_data:
                return "No data available for analysis."
            
            # Calculate diversity metrics
            control_diversity = self._calculate_shannon_diversity([t['control_avg'] for t in taxa_data])
            uc_diversity = self._calculate_shannon_diversity([t['uc_avg'] for t in taxa_data])
            
            # Find top 20 most distinct taxa
            top_taxa = sorted(taxa_data, key=lambda x: abs(x['control_uc_ratio'] - 1) if x['control_uc_ratio'] != float('inf') else 0, reverse=True)[:20]
            
            # Prepare data for AI analysis
            analysis_data = {
                'taxonomic_level': taxonomic_level,
                'total_taxa': len(taxa_data),
                'control_diversity': control_diversity,
                'uc_diversity': uc_diversity,
                'top_taxa': top_taxa,
                'report_type': report_type
            }
            
            # Generate AI prompt
            prompt = self._create_summary_prompt(analysis_data)
            
            # Call AI model
            summary = self._call_ollama_summary(prompt)
            
            return summary
            
        except Exception as e:
            print(f"Error generating AI summary: {e}")
            # Return a helpful error message with setup instructions
            return f"""
            **Error Generating AI Summary:**
            
            An error occurred while generating the AI summary: {e}
            
            **Troubleshooting:**
            1. Ensure Ollama is installed and running
            2. Install the gpt-oss:20b model: `ollama pull gpt-oss:20b`
            3. Check if Ollama is accessible at http://localhost:11434
            
            **Alternative:**
            The system will generate a fallback summary with basic analysis.
            """
    
    def _calculate_shannon_diversity(self, abundances: List[float]) -> float:
        """Calculate Shannon diversity index"""
        try:
            # Remove zero abundances
            non_zero = [x for x in abundances if x > 0]
            if not non_zero:
                return 0.0
            
            total = sum(non_zero)
            proportions = [x/total for x in non_zero]
            
            # Calculate Shannon index
            shannon = -sum(p * np.log(p) for p in proportions if p > 0)
            return shannon
        except:
            return 0.0
    
    def _create_summary_prompt(self, data: Dict[str, Any]) -> str:
        """Create AI prompt for summary generation"""
        
        if data['report_type'] == 'lay':
            style_instruction = """
            Write in simple, easy-to-understand language suitable for non-scientists. 
            Use everyday analogies and avoid technical jargon. 
            Focus on what the results mean for health and well-being.
            """
        else:
            style_instruction = """
            Write in scientific, technical language suitable for researchers and clinicians. 
            Use precise terminology and include statistical interpretations.
            Focus on biological significance and research implications.
            """
        
        # Create top taxa table for prompt
        taxa_table = "Top 20 Most Distinct Taxa:\n"
        taxa_table += "Taxon | Control_Avg | UC_Avg | UC_StdDev | Ratio\n"
        taxa_table += "-" * 60 + "\n"
        
        for i, taxon in enumerate(data['top_taxa'][:20], 1):
            ratio_str = f"{taxon['control_uc_ratio']:.2f}" if taxon['control_uc_ratio'] != float('inf') else "Inf"
            taxa_table += f"{i}. {taxon['taxon']} | {taxon['control_avg']:.4f} | {taxon['uc_avg']:.4f} | {taxon['uc_std']:.4f} | {ratio_str}\n"
        
        prompt = f"""
        You are a microbiome analysis expert. Generate a comprehensive summary report based on the following data.
        
        {style_instruction}
        
        ANALYSIS REQUIREMENTS:
        - Word limit: 1000 words
        - Include diversity analysis comparing Control vs UC samples
        - Analyze the largest positive and negative ratios
        - Interpret which taxa are increased or decreased in UC samples
        - Provide biological significance and health implications
        
        DATA SUMMARY:
        - Taxonomic Level: {data['taxonomic_level']}
        - Total Taxa Analyzed: {data['total_taxa']}
        - Control Sample Shannon Diversity: {data['control_diversity']:.4f}
        - UC Sample Shannon Diversity: {data['uc_diversity']:.4f}
        
        {taxa_table}
        
        Please provide a well-structured report with:
        1. Executive Summary
        2. Diversity Analysis
        3. Key Taxa Changes
        4. Biological Interpretation
        5. Health Implications
        6. Recommendations for Further Analysis
        
        Focus on the most biologically significant findings and their implications.
        """
        
        return prompt
    
    def _call_ollama_summary(self, prompt: str) -> str:
        """Call Ollama API for summary generation"""
        try:
            import requests
            
            # Ollama API endpoint
            url = "http://localhost:11434/api/generate"
            
            # Request payload
            payload = {
                "model": "gpt-oss:20b",
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.7,
                    "top_p": 0.9,
                    "max_tokens": 1500
                }
            }
            
            # Make request
            response = requests.post(url, json=payload, timeout=120)
            
            if response.status_code == 200:
                result = response.json()
                return result.get('response', 'No response generated')
            else:
                return f"Error calling Ollama API: {response.status_code}"
                
        except Exception as e:
            print(f"Error calling Ollama: {e}")
            # Generate a fallback summary when Ollama is not available
            return self._generate_fallback_summary(prompt)
    
    def _generate_fallback_summary(self, prompt: str) -> str:
        """Generate a fallback summary when Ollama is not available"""
        try:
            # Extract key information from the prompt
            if "family" in prompt.lower():
                level = "Family"
            elif "genus" in prompt.lower():
                level = "Genus"
            elif "species" in prompt.lower():
                level = "Species"
            elif "phylum" in prompt.lower():
                level = "Phylum"
            elif "class" in prompt.lower():
                level = "Class"
            elif "order" in prompt.lower():
                level = "Order"
            else:
                level = "Taxonomic"
            
            # Generate a basic summary
            summary = f"""
            **AI-Powered Microbiome Analysis Report - {level} Level**
            
            **Executive Summary:**
            This report provides a comprehensive analysis of microbiome data at the {level.lower()} level, comparing Control and UC (Ulcerative Colitis) samples.
            
            **Diversity Analysis:**
            The analysis reveals differences in microbial diversity between Control and UC samples. Shannon diversity indices are calculated to quantify these differences.
            
            **Key Findings:**
            - Control samples show distinct microbial composition compared to UC samples
            - Several taxa demonstrate significant abundance differences between groups
            - The analysis identifies the top 20 most distinct taxa between Control and UC samples
            
            **Biological Interpretation:**
            Changes in microbial composition may reflect the inflammatory environment in UC patients. Reduced diversity in UC samples suggests a less stable microbiome ecosystem.
            
            **Health Implications:**
            Understanding these microbial differences can help develop targeted interventions and provide insights into disease mechanisms.
            
            **Recommendations:**
            - Further analysis with larger sample sizes
            - Functional analysis of identified taxa
            - Longitudinal studies to track changes over time
            
            **Note:** This is a fallback summary generated when the AI service is unavailable. For enhanced AI-powered analysis, please ensure Ollama is running with the gpt-oss:20b model.
            """
            
            return summary
            
        except Exception as e:
            return f"Error generating fallback summary: {e}"
    
    def generate_diversity_plot(self, taxonomic_level: str) -> str:
        """Generate diversity comparison plot data"""
        try:
            taxa_data = self.calculate_taxa_control_uc_ratios(taxonomic_level)
            
            if not taxa_data:
                return ""
            
            # Calculate diversity for different sample sizes
            control_abundances = [t['control_avg'] for t in taxa_data]
            uc_abundances = [t['uc_avg'] for t in taxa_data]
            
            # Create plot data
            plot_data = {
                'labels': ['Control', 'UC'],
                'diversity_values': [
                    self._calculate_shannon_diversity(control_abundances),
                    self._calculate_shannon_diversity(uc_abundances)
                ],
                'sample_counts': [len([x for x in control_abundances if x > 0]), 
                                len([x for x in uc_abundances if x > 0])]
            }
            
            return plot_data
            
        except Exception as e:
            print(f"Error generating diversity plot: {e}")
            return ""
    
    def generate_heatmap_data(self, taxonomic_level: str) -> Dict[str, Any]:
        """Generate heatmap data for top 20 most distinct taxa"""
        try:
            taxa_data = self.calculate_taxa_control_uc_ratios(taxonomic_level)
            
            if not taxa_data:
                return {}
            
            # Get top 20 most distinct taxa
            top_taxa = sorted(taxa_data, 
                             key=lambda x: abs(x['control_uc_ratio'] - 1) if x['control_uc_ratio'] != float('inf') else 0, 
                             reverse=True)[:20]
            
            # Prepare heatmap data
            heatmap_data = {
                'taxa': [t['taxon'] for t in top_taxa],
                'control_values': [t['control_avg'] for t in top_taxa],
                'uc_values': [t['uc_avg'] for t in top_taxa],
                'ratios': [t['control_uc_ratio'] if t['control_uc_ratio'] != float('inf') else 10.0 for t in top_taxa]
            }
            
            return heatmap_data
            
        except Exception as e:
            print(f"Error generating heatmap data: {e}")
            return {}

def main():
    """
    Test the real-time analyzer
    """
    analyzer = RealTimeMicrobiomeAnalyzer()
    
    # Test data
    test_data = {
        "PedCtrl59": 0.01,
        "SRR15702647": 0.23,
        "SRR15702658": 0.38,
        "SRR15702659": 0.31,
        "SRR15702660": 0.29
    }
    
    sample_names = ["PedCtrl59", "SRR15702647", "SRR15702658", "SRR15702659", "SRR15702660"]
    rpm_values = [0.01, 0.23, 0.38, 0.31, 0.29]
    
    print("Testing Real-Time AI Analyzer...")
    print("=" * 50)
    
    try:
        result = analyzer.analyze_taxon_plot("Acetobacteraceae", test_data, sample_names, rpm_values)
        print("AI Analysis Result:")
        print(result)
    except Exception as e:
        print(f"Error: {e}")
        print("Fallback analysis:")
        fallback = analyzer._generate_fallback_analysis("Acetobacteraceae", test_data, sample_names, rpm_values)
        print(fallback)
    
    # Test diversity data extraction
    print("\n" + "=" * 50)
    print("Testing Diversity Data Extraction...")
    
    diversity_data = analyzer.extract_alpha_diversity_data("phylum")
    if diversity_data:
        print("Alpha diversity data extracted:")
        for sample, diversity in diversity_data.items():
            print(f"  {sample}: {diversity:.3f}")
        
        print("\nGenerating AI analysis for diversity...")
        try:
            result = analyzer.analyze_diversity_plot("alpha_diversity", diversity_data)
            print("AI Diversity Analysis:")
            print(result)
        except Exception as e:
            print(f"Error: {e}")
            fallback = analyzer._generate_fallback_diversity_analysis("alpha_diversity", diversity_data)
            print("Fallback diversity analysis:")
            print(fallback)
    else:
        print("No diversity data found")

if __name__ == "__main__":
    main()
